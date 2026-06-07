#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HCS_Dynamic.py - 计算肽段的HCS动态指标

修复说明（相对于V3）：
1. 区分HMom_HoF和HMom_HoF_HELNET：
   - HMom_HoF：直接使用疏水面核心位置（HoF_AA_Num）按螺旋轮空间角度计算HMom
   - HMom_HoF_HELNET：使用疏水面位置及其±1,±3,±4邻居中疏水性且角度差≤100°的位置计算HMom（原HMom_HoF逻辑）
   - 新增HoF_AA_HELNET_Num：记录HMom_HoF_HELNET实际使用的氨基酸编号
2. 所有HCS指标（HCS4_HoF_Mean, HCS4_HoF_Mean_SD, HCS4_HoF_Num, HCS3_HoF_Mean, HCS3_HoF_Mean_SD, HCS3_HoF_Num, HCS_Pair_Num及对应的HCS*_i列）
   均直接基于疏水面核心位置（HoF_AA_Num），不再扩展邻居，确保计算符合螺旋轮空间角度逻辑。

功能说明：
1. 读取包含Name, Sequences, Exp_Log2_MIC的txt文件
2. 根据α螺旋空间结构原理计算残基角度 = (i-1)*100° mod 360
3. 实现平均疏水矩HMom方法（基于Eisenberg公式）
4. 计算疏水面相关指标：
   - HMom_HoF: 疏水面核心疏水矩
   - HMom_HoF_HELNET: 疏水面扩展邻居疏水矩
   - HCS4_HoF_Mean, HCS4_HoF_Mean_SD, HCS4_HoF_Num: i+4成对关系
   - HCS3_HoF_Mean, HCS3_HoF_Mean_SD, HCS3_HoF_Num: i+3成对关系
   - HCS_Pair_Num, HCS_Pair*: i, i+7或i+3, i+7三重关系
5. 输出xlsx表格，包含HydrophobicFace, HoF_AA_Num, HMom_HoF, HMom_HoF_HELNET, HoF_AA_HELNET_Num等字段

使用方法：
python HCS_Dynamic.py -i Q1_19_Seq.txt -o Q1_19_Seq_HCS_V5.xlsx
"""

import os
import sys
import math
import argparse
from typing import List, Tuple, Dict, Set, Optional
from collections import Counter

# 尝试导入openpyxl用于输出Excel
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Please install with: pip install openpyxl")

# 添加当前目录到路径，以便导入PeptidePipeline_Core_V3
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# 尝试导入PeptidePipeline_Core_V3的函数
PEPTIDE_PIPELINE_AVAILABLE = False
try:
    from PeptidePipeline_Core_V3 import (
        find_hydrophobic_face_multiround,
        find_helical_faces_multiround,
        FAUCHERE_PLISKA,
        HYDROPHOBIC_FACE_RESIDUES,
        ROUND_SIZE
    )
    PEPTIDE_PIPELINE_AVAILABLE = True
except ImportError:
    print("Warning: PeptidePipeline_Core_V3.py not found, using fallback hydrophobic face detection")

# ============================ 常量定义 ============================

# Fauchère-Pliska疏水性 scale (1983)
DEFAULT_FAUCHERE_PLISKA = {
    'A': 0.310, 'R': -1.010, 'N': -0.600, 'D': -0.770, 'C': 1.540,
    'Q': -0.220, 'E': -0.640, 'G': 0.000, 'H': 0.130, 'I': 1.800,
    'L': 1.700, 'K': -0.990, 'M': 1.230, 'F': 1.790, 'P': 0.720,
    'S': -0.040, 'T': 0.260, 'W': 2.250, 'Y': 0.960, 'V': 1.220
}

# 疏水性氨基酸集合
HYDROPHOBIC = {'A', 'L', 'I', 'V', 'M', 'P', 'F', 'W', 'Y'}

# 螺旋轮参数
DEGREES_PER_RESIDUE = 100.0  # 每残基旋转角度
DELTA_RADIANS = math.radians(DEGREES_PER_RESIDUE)  # 100度对应的弧度

# ============================ 辅助函数 ============================

def get_fauchere_pliska() -> Dict[str, float]:
    """获取疏水性 scale"""
    return FAUCHERE_PLISKA if PEPTIDE_PIPELINE_AVAILABLE else DEFAULT_FAUCHERE_PLISKA

def calculate_helix_angle(position: int) -> float:
    """
    计算指定位置氨基酸的螺旋角度
    公式: angle = (position - 1) * 100° mod 360

    Args:
        position: 氨基酸编号（1-based）

    Returns:
        角度值（0-360度）
    """
    return ((position - 1) * DEGREES_PER_RESIDUE) % 360.0

def calculate_angle_difference(angle1: float, angle2: float) -> float:
    """
    计算两个角度之间的最小差值（考虑周期性）
    返回值范围: [0, 180]
    """
    diff = abs(angle1 - angle2)
    if diff > 180:
        diff = 360 - diff
    return diff

def is_hydrophobic_amino_acid(aa: str) -> bool:
    """判断氨基酸是否为疏水性"""
    return aa in HYDROPHOBIC

def get_neighbor_positions(position: int, offsets: List[int] = None) -> List[int]:
    """
    获取指定位置的邻居位置（±1, ±3, ±4）

    Args:
        position: 基础位置
        offsets: 偏移列表，默认[-4, -3, -1, 1, 3, 4]

    Returns:
        邻居位置列表（位置编号大于0的）
    """
    if offsets is None:
        offsets = [-4, -3, -1, 1, 3, 4]

    neighbors = []
    for offset in offsets:
        neighbor = position + offset
        if neighbor > 0:  # 位置编号必须为正整数
            neighbors.append(neighbor)
    return neighbors

# ============================ HMom计算 ============================

def calculate_mean_amphipathic_moment(sequence: str) -> float:
    """
    计算给定序列的平均疏水矩（Hydrophobic Moment）

    公式（Eisenberg 1982）：
    HMom = (1/N) * sqrt(sum(H_n * sin(n*delta))^2 + sum(H_n * cos(n*delta))^2)

    其中：
    - N: 氨基酸数量
    - H_n: 第n个氨基酸的疏水性值（Fauchère-Pliska scale）
    - delta: 相邻残基间的旋转角度（alpha螺旋为100度 = 100*pi/180弧度）

    Args:
        sequence: 氨基酸序列

    Returns:
        平均疏水矩值
    """
    if not sequence:
        return 0.0

    fauchere_pliska = get_fauchere_pliska()
    n = len(sequence)

    sum_sin = 0.0
    sum_cos = 0.0

    for i, aa in enumerate(sequence):
        hydrophobicity = fauchere_pliska.get(aa, 0.0)
        angle = i * DELTA_RADIANS
        sum_sin += hydrophobicity * math.sin(angle)
        sum_cos += hydrophobicity * math.cos(angle)

    hmom = math.sqrt(sum_sin * sum_sin + sum_cos * sum_cos) / n
    return hmom

def calculate_hmom_for_positions_direct(positions: List[int], sequence: str) -> float:
    """
    直接根据指定位置列表（排序后）提取子序列并计算HMom，不扩展邻居

    Args:
        positions: 位置列表（1-based）
        sequence: 完整序列

    Returns:
        HMom值
    """
    if not positions:
        return 0.0

    sorted_positions = sorted(positions)
    sub_seq = ''.join(sequence[pos - 1] for pos in sorted_positions if 1 <= pos <= len(sequence))
    return calculate_mean_amphipathic_moment(sub_seq)

def calculate_hmom_for_positions_with_neighbors_and_positions(
    sequence: str,
    positions: List[int],
    pos_to_angle: Dict[int, float],
    pos_to_aa: Dict[int, str]
) -> Tuple[float, List[int]]:
    """
    计算指定位置集合的平均疏水矩（包括其±1,±3,±4邻居中符合条件的疏水性氨基酸）
    并返回实际使用的所有位置（1-based，排序后）

    计算方法：
    1. 对每个疏水面位置，计算其±1, ±3, ±4邻居
    2. 筛选邻居：只保留疏水性氨基酸，且角度差<=100度
    3. 收集所有符合条件的邻居位置（包括原位置）
    4. 使用HMom公式计算这些位置的平均疏水矩

    Args:
        sequence: 氨基酸序列
        positions: 疏水面位置列表（1-based）
        pos_to_angle: 位置到角度的映射
        pos_to_aa: 位置到氨基酸的映射

    Returns:
        (hmom, used_positions): 平均疏水矩值和实际使用的位置列表（排序后）
    """
    if not positions:
        return 0.0, []

    n = len(sequence)
    valid_positions: Set[int] = set()

    for pos in positions:
        if pos not in pos_to_angle:
            continue

        base_angle = pos_to_angle[pos]

        # 首先将疏水面位置本身加入
        if pos in pos_to_aa and is_hydrophobic_amino_acid(pos_to_aa[pos]):
            valid_positions.add(pos)

        # 获取±1, ±3, ±4的邻居位置
        neighbors = get_neighbor_positions(pos)

        for neighbor in neighbors:
            if neighbor < 1 or neighbor > n:
                continue
            if neighbor not in pos_to_aa:
                continue

            neighbor_aa = pos_to_aa[neighbor]
            if not is_hydrophobic_amino_acid(neighbor_aa):
                continue

            if neighbor in pos_to_angle:
                neighbor_angle = pos_to_angle[neighbor]
                angle_diff = calculate_angle_difference(base_angle, neighbor_angle)
                if angle_diff <= 100:
                    valid_positions.add(neighbor)

    if len(valid_positions) < 2:
        # 不足两个位置，直接计算单个位置的HMom（但单个位置定义可能为0）
        if len(valid_positions) == 1:
            pos = next(iter(valid_positions))
            sub_seq = sequence[pos - 1]
            return calculate_mean_amphipathic_moment(sub_seq), sorted(valid_positions)
        return 0.0, []

    sorted_positions = sorted(valid_positions)
    sub_seq = ''.join(sequence[pos - 1] for pos in sorted_positions)
    hmom = calculate_mean_amphipathic_moment(sub_seq)
    return hmom, sorted_positions

def calculate_hmom_for_pair_direct(pair: Tuple[int, ...], sequence: str) -> float:
    """
    直接计算一对或一组位置的HMom值（不扩展邻居）

    Args:
        pair: 位置元组（2个或3个位置）
        sequence: 氨基酸序列

    Returns:
        HMom值
    """
    if len(pair) < 2:
        return 0.0
    positions = sorted(pair)
    sub_seq = ''.join(sequence[pos - 1] for pos in positions if 1 <= pos <= len(sequence))
    return calculate_mean_amphipathic_moment(sub_seq)

def calculate_standard_deviation(values: List[float]) -> float:
    """
    计算标准差（使用无偏估计量 n-1）

    Args:
        values: 数值列表

    Returns:
        标准差
    """
    if len(values) < 2:
        return 0.0

    n = len(values)
    mean = sum(values) / n
    variance = sum((x - mean) ** 2 for x in values) / (n - 1)  # 无偏估计
    return math.sqrt(variance)

# ============================ 疏水面检测 ============================

def find_hydrophobic_face_positions(sequence: str) -> Tuple[List[int], str]:
    """
    查找疏水面氨基酸位置

    从PeptidePipeline_Core_V3调用find_hydrophobic_face_multiround获取疏水面信息，
    返回1-based的氨基酸位置编号

    处理规则:
    - 如果疏水面最后一个氨基酸为G，则移除G，只保留前面的疏水面氨基酸
    - 例如: 疏水面为FLLLMG，则只保留FLLLM

    Args:
        sequence: 氨基酸序列

    Returns:
        (positions, face_sequence): 疏水面位置列表（1-based）和氨基酸序列
    """
    if PEPTIDE_PIPELINE_AVAILABLE:
        # 调用PeptidePipeline_Core_V3的函数
        _, _, hydrophobic_face_seq, hyd_linear_idxs = find_hydrophobic_face_multiround(sequence)

        if not hyd_linear_idxs:
            return [], ""

        # 转换为1-based编号
        positions = [idx + 1 for idx in hyd_linear_idxs]

        # 获取疏水面序列
        face_seq = hydrophobic_face_seq

    else:
        # Fallback: 如果无法导入PeptidePipeline，使用简单的疏水性判断
        positions = []
        face_seq = ""

        for i, aa in enumerate(sequence):
            if aa in HYDROPHOBIC:
                positions.append(i + 1)  # 1-based
                face_seq += aa

    # 如果疏水面最后一个氨基酸为G，则移除
    if face_seq and face_seq[-1] == 'G':
        face_seq = face_seq[:-1]
        if positions:
            positions = positions[:-1]

    return positions, face_seq

# ============================ HCS成对和三重关系 ============================

def find_hcs4_pairs(positions: List[int]) -> List[Tuple[int, int]]:
    """
    查找i+4成对关系

    Args:
        positions: 疏水面位置列表（1-based，无需排序）

    Returns:
        成对关系列表，每对位置已排序
    """
    position_set = set(positions)
    pairs = []

    for pos in positions:
        partner = pos + 4
        if partner in position_set:
            pairs.append(tuple(sorted((pos, partner))))

    # 去重（基于set，因为每对可能出现两次）
    unique_pairs = list(set(pairs))
    return unique_pairs

def find_hcs3_pairs(positions: List[int]) -> List[Tuple[int, int]]:
    """
    查找i+3成对关系

    Args:
        positions: 疏水面位置列表（1-based）

    Returns:
        成对关系列表，每对位置已排序
    """
    position_set = set(positions)
    pairs = []

    for pos in positions:
        partner = pos + 3
        if partner in position_set:
            pairs.append(tuple(sorted((pos, partner))))

    unique_pairs = list(set(pairs))
    return unique_pairs

def find_hcs_pairs_triplet(positions: List[int]) -> List[Tuple[int, int, int]]:
    """
    查找(i, i+3, i+7)或(i, i+4, i+7)三重关系

    逻辑：
    1. 对于每个位置i，检查是否存在i+3和i+7（都在疏水面中）
    2. 对于每个位置i，检查是否存在i+4和i+7（都在疏水面中）

    Args:
        positions: 疏水面位置列表（1-based）

    Returns:
        三重关系列表，每个三元组已排序
    """
    position_set = set(positions)
    triplets_set = set()

    # 方式1: (i, i+3, i+7)
    for pos in positions:
        if pos + 3 in position_set and pos + 7 in position_set:
            triplet = tuple(sorted((pos, pos + 3, pos + 7)))
            triplets_set.add(triplet)

    # 方式2: (i, i+4, i+7)
    for pos in positions:
        if pos + 4 in position_set and pos + 7 in position_set:
            triplet = tuple(sorted((pos, pos + 4, pos + 7)))
            triplets_set.add(triplet)

    # 转换为列表并排序（保证输出稳定）
    triplets = sorted(triplets_set)
    return triplets

# ============================ 主计算函数 ============================

def calculate_hcs_features(sequence: str, name: str = "") -> Dict:
    """
    计算一条序列的所有HCS特征（V5版本）

    Args:
        sequence: 氨基酸序列
        name: 序列名称（可选）

    Returns:
        包含所有特征值的字典
    """
    n = len(sequence)

    if n == 0:
        return {
            'Name': name,
            'Sequences': sequence,
            'Exp_Log2_MIC': None,
            'HMom': 0.0,
            'HydrophobicFace': '',
            'HoF_AA_Num': '',
            'HMom_HoF': 0.0,
            'HMom_HoF_HELNET': 0.0,
            'HoF_AA_HELNET_Num': '',
            'HCS4_HoF_Mean': 0.0,
            'HCS4_HoF_Mean_SD': 0.0,
            'HCS4_HoF_Num': 0,
            'HCS3_HoF_Mean': 0.0,
            'HCS3_HoF_Mean_SD': 0.0,
            'HCS3_HoF_Num': 0,
            'HCS_Pair_Num': 0
        }

    # 构建位置到角度和氨基酸的映射
    pos_to_angle = {}
    pos_to_aa = {}

    for i in range(n):
        position = i + 1
        aa = sequence[i]
        angle = calculate_helix_angle(position)

        pos_to_angle[position] = angle
        pos_to_aa[position] = aa

    # 计算整个序列的HMom
    hmom_total = calculate_mean_amphipathic_moment(sequence)

    # 查找疏水面位置（1-based）
    hof_positions, hof_seq = find_hydrophobic_face_positions(sequence)

    # 格式化输出疏水面位置编号（排序后）
    hof_seq_str = hof_seq if hof_seq else ""
    hof_nums_str = "-".join(map(str, sorted(hof_positions))) if hof_positions else ""

    if not hof_positions:
        # 没有疏水面，返回默认值
        return {
            'Name': name,
            'Sequences': sequence,
            'Exp_Log2_MIC': None,
            'HMom': hmom_total,
            'HydrophobicFace': hof_seq_str,
            'HoF_AA_Num': hof_nums_str,
            'HMom_HoF': 0.0,
            'HMom_HoF_HELNET': 0.0,
            'HoF_AA_HELNET_Num': '',
            'HCS4_HoF_Mean': 0.0,
            'HCS4_HoF_Mean_SD': 0.0,
            'HCS4_HoF_Num': 0,
            'HCS3_HoF_Mean': 0.0,
            'HCS3_HoF_Mean_SD': 0.0,
            'HCS3_HoF_Num': 0,
            'HCS_Pair_Num': 0
        }

    # 1. HMom_HoF: 直接使用疏水面核心位置计算HMom
    HMom_HoF = calculate_hmom_for_positions_direct(hof_positions, sequence)

    # 2. HMom_HoF_HELNET 和 HoF_AA_HELNET_Num: 扩展邻居后计算
    hmom_hf_helnet, used_positions_helnet = calculate_hmom_for_positions_with_neighbors_and_positions(
        sequence, hof_positions, pos_to_angle, pos_to_aa
    )
    helnet_nums_str = "-".join(map(str, used_positions_helnet)) if used_positions_helnet else ""

    # 3. HCS4: i+4成对关系（基于疏水面核心位置，不扩展邻居）
    hcs4_pairs = find_hcs4_pairs(hof_positions)
    hcs4_num = len(hcs4_pairs)
    hcs4_values = []
    for pair in hcs4_pairs:
        pair_hmom = calculate_hmom_for_pair_direct(pair, sequence)
        hcs4_values.append(pair_hmom)

    if hcs4_values:
        hcs4_mean = sum(hcs4_values) / len(hcs4_values)
        hcs4_sd = calculate_standard_deviation(hcs4_values)
    else:
        hcs4_mean = 0.0
        hcs4_sd = 0.0

    # 4. HCS3: i+3成对关系
    hcs3_pairs = find_hcs3_pairs(hof_positions)
    hcs3_num = len(hcs3_pairs)
    hcs3_values = []
    for pair in hcs3_pairs:
        pair_hmom = calculate_hmom_for_pair_direct(pair, sequence)
        hcs3_values.append(pair_hmom)

    if hcs3_values:
        hcs3_mean = sum(hcs3_values) / len(hcs3_values)
        hcs3_sd = calculate_standard_deviation(hcs3_values)
    else:
        hcs3_mean = 0.0
        hcs3_sd = 0.0

    # 5. HCS_Pair: 三重关系
    hcs_triplets = find_hcs_pairs_triplet(hof_positions)
    hcs_pair_num = len(hcs_triplets)
    hcs_pair_values = []
    for triplet in hcs_triplets:
        triplet_hmom = calculate_hmom_for_pair_direct(triplet, sequence)
        hcs_pair_values.append(triplet_hmom)

    # 构建结果字典
    result = {
        'Name': name,
        'Sequences': sequence,
        'Exp_Log2_MIC': None,
        'HMom': hmom_total,
        'HydrophobicFace': hof_seq_str,
        'HoF_AA_Num': hof_nums_str,
        'HMom_HoF': HMom_HoF,
        'HMom_HoF_HELNET': hmom_hf_helnet,
        'HoF_AA_HELNET_Num': helnet_nums_str,
        'HCS4_HoF_Mean': hcs4_mean,
        'HCS4_HoF_Mean_SD': hcs4_sd,
        'HCS4_HoF_Num': hcs4_num,
        'HCS3_HoF_Mean': hcs3_mean,
        'HCS3_HoF_Mean_SD': hcs3_sd,
        'HCS3_HoF_Num': hcs3_num,
        'HCS_Pair_Num': hcs_pair_num
    }

    # 添加 HCS4_HoF_i 列
    for i, value in enumerate(hcs4_values, 1):
        result[f'HCS4_HoF_{i}'] = value

    # 添加 HCS3_HoF_i 列
    for i, value in enumerate(hcs3_values, 1):
        result[f'HCS3_HoF_{i}'] = value

    # 添加 HCS_Pair_i 列
    for i, value in enumerate(hcs_pair_values, 1):
        result[f'HCS_Pair{i}'] = value

    return result

# ============================ 文件处理 ============================

def read_input_file(input_path: str) -> List[Dict]:
    """
    读取输入文件

    文件格式: tab分隔，包含Name, Sequences, Exp_Log2_MIC列

    Args:
        input_path: 输入文件路径

    Returns:
        记录列表
    """
    records = []

    try:
        with open(input_path, 'r', encoding='utf-8') as f:
            first_line = f.readline().strip()

            if '\t' in first_line:
                delimiter = '\t'
            else:
                delimiter = ','

            headers = first_line.split(delimiter)
            headers = [h.strip() for h in headers]

            for line in f:
                line = line.strip()
                if not line:
                    continue

                fields = line.split(delimiter)
                fields = [f.strip() for f in fields]

                record = {}
                for i, header in enumerate(headers):
                    if i < len(fields):
                        record[header] = fields[i]
                    else:
                        record[header] = ''

                records.append(record)

    except FileNotFoundError:
        print(f"Error: File '{input_path}' not found")
        sys.exit(1)
    except Exception as e:
        print(f"Error reading file: {e}")
        sys.exit(1)

    return records

def write_output_file(records: List[Dict], output_path: str):
    """
    写入输出文件（Excel格式）

    Args:
        records: 记录列表
        output_path: 输出文件路径
    """
    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl is not available. Cannot write Excel file.")
        print("Please install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HCS_Results_V5"

    if not records:
        print("No records to write")
        return

    all_columns = set()
    for record in records:
        all_columns.update(record.keys())

    # 定义列顺序（优先显示核心列）
    priority_columns = [
        'Name', 'Sequences', 'Exp_Log2_MIC', 'HMom',
        'HydrophobicFace', 'HoF_AA_Num',
        'HMom_HoF', 'HMom_HoF_HELNET', 'HoF_AA_HELNET_Num',
        'HCS4_HoF_Mean', 'HCS4_HoF_Mean_SD', 'HCS4_HoF_Num',
        'HCS3_HoF_Mean', 'HCS3_HoF_Mean_SD', 'HCS3_HoF_Num',
        'HCS_Pair_Num'
    ]

    # 收集动态列
    dynamic_columns = []
    for record in records:
        for key in record.keys():
            if key not in priority_columns and key not in dynamic_columns:
                dynamic_columns.append(key)

    # 过滤并排序动态列（HCS4_HoF_*, HCS3_HoF_*, HCS_Pair*）
    def dynamic_sort_key(col):
        if col.startswith('HCS4_HoF_'):
            parts = col.split('_')
            if len(parts) == 3 and parts[2].isdigit():
                return (1, int(parts[2]))
            return (1, 0)
        elif col.startswith('HCS3_HoF_'):
            parts = col.split('_')
            if len(parts) == 3 and parts[2].isdigit():
                return (2, int(parts[2]))
            return (2, 0)
        elif col.startswith('HCS_Pair'):
            if col[8:].isdigit():
                return (3, int(col[8:]))
            return (3, 0)
        else:
            return (4, 0)

    dynamic_columns.sort(key=dynamic_sort_key)

    # 构建最终列顺序
    column_order = []
    for col in priority_columns:
        if col in all_columns:
            column_order.append(col)
    for col in dynamic_columns:
        if col in all_columns and col not in column_order:
            column_order.append(col)

    # 写入表头
    for col_idx, col_name in enumerate(column_order, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name

    # 写入数据
    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(column_order, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = record.get(col_name, '')
            cell.value = value

    # 自动调整列宽
    for col_idx in range(1, len(column_order) + 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        max_length = 0

        for row_idx in range(1, len(records) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Output saved to: {output_path}")

# ============================ 主函数 ============================

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='HCS_Dynamic.py - 计算肽段的HCS动态指标',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例用法:
  python HCS_Dynamic.py -i input.txt -o results_V5.xlsx

输入文件格式 (tab分隔):
  Name    Sequences    Exp_Log2_MIC
  Peptide1    MRKLLKKLHRFKAKLVRGG    2.5
  Peptide2    AALAAKAAAKLAAA    3.0

主要改进（V5）:
  - 区分 HMom_HoF（疏水面核心）和 HMom_HoF_HELNET（扩展邻居）
  - 所有 HCS 指标均基于疏水面核心位置，符合螺旋轮空间角度逻辑
  - 新增 HoF_AA_HELNET_Num 记录实际使用的氨基酸编号
        """
    )

    parser.add_argument('-i', '--input', required=True, help='输入文件路径')
    parser.add_argument('-o', '--output', required=True, help='输出Excel文件路径')
    parser.add_argument('-v', '--verbose', action='store_true', help='显示详细输出')

    args = parser.parse_args()

    print(f"Reading input file: {args.input}")
    records = read_input_file(args.input)
    print(f"Loaded {len(records)} records")

    results = []

    for idx, record in enumerate(records):
        name = record.get('Name', f'Sequence_{idx + 1}')
        sequence = record.get('Sequences', '')
        exp_log2_mic = record.get('Exp_Log2_MIC', '')

        if args.verbose:
            print(f"Processing: {name} - {sequence[:20]}...")

        features = calculate_hcs_features(sequence, name)
        features['Exp_Log2_MIC'] = exp_log2_mic

        results.append(features)

        if args.verbose:
            print(f"  HydrophobicFace: {features['HydrophobicFace']}")
            print(f"  HoF_AA_Num: {features['HoF_AA_Num']}")
            print(f"  HMom_HoF: {features['HMom_HoF']:.4f}")
            print(f"  HMom_HoF_HELNET: {features['HMom_HoF_HELNET']:.4f}")
            print(f"  HoF_AA_HELNET_Num: {features['HoF_AA_HELNET_Num']}")
            print(f"  HCS4_HoF_Mean: {features['HCS4_HoF_Mean']:.4f}, Num: {features['HCS4_HoF_Num']}")
            print(f"  HCS3_HoF_Mean: {features['HCS3_HoF_Mean']:.4f}, Num: {features['HCS3_HoF_Num']}")
            print(f"  HCS_Pair_Num: {features['HCS_Pair_Num']}")

    print(f"Writing output file: {args.output}")
    write_output_file(results, args.output)

    print("Done!")

if __name__ == '__main__':
    main()